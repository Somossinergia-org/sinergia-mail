#!/usr/bin/env python3
"""Generate Excel reports for Sinergia Mail dashboard."""

import sys
import json
from datetime import datetime
from pathlib import Path
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def format_currency(value):
    """Format value as currency."""
    if value is None:
        return 0
    return float(value)


def add_header_row(ws, headers):
    """Add formatted header row to worksheet."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_alternating_rows(ws, start_row, end_row, start_col=1, end_col=None):
    """Add alternating row colors."""
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row_num in range(start_row, end_row + 1):
        if row_num % 2 == 0:
            for col_num in range(start_col, (end_col or ws.max_column) + 1):
                ws.cell(row=row_num, column=col_num).fill = light_fill


def add_thin_border(ws, start_row, end_row, start_col=1, end_col=None):
    """Add thin borders to range."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_num in range(start_row, end_row + 1):
        for col_num in range(start_col, (end_col or ws.max_column) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border


def generate_invoices_report(data):
    """Generate invoices report with Facturas and Resumen sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Facturas (Invoices)
    ws_invoices = wb.create_sheet("Facturas", 0)
    headers = ["Nº Factura", "Emisor", "NIF", "Concepto", "Base Imponible", "IVA", "Total", "Moneda", "Fecha", "Categoría"]
    add_header_row(ws_invoices, headers)

    invoices = data.get("invoices", [])
    for idx, invoice in enumerate(invoices, start=2):
        ws_invoices.cell(row=idx, column=1).value = invoice.get("invoiceNumber", "")
        ws_invoices.cell(row=idx, column=2).value = invoice.get("issuerName", "")
        ws_invoices.cell(row=idx, column=3).value = invoice.get("issuerNif", "")
        ws_invoices.cell(row=idx, column=4).value = invoice.get("concept", "")

        # Amount (Base Imponible)
        amount_cell = ws_invoices.cell(row=idx, column=5)
        amount_cell.value = format_currency(invoice.get("amount"))
        amount_cell.number_format = "#,##0.00"

        # Tax (IVA)
        tax_cell = ws_invoices.cell(row=idx, column=6)
        tax_cell.value = format_currency(invoice.get("tax"))
        tax_cell.number_format = "#,##0.00"

        # Total
        total_cell = ws_invoices.cell(row=idx, column=7)
        total_cell.value = format_currency(invoice.get("totalAmount"))
        total_cell.number_format = "#,##0.00"

        ws_invoices.cell(row=idx, column=8).value = invoice.get("currency", "EUR")

        # Date
        date_val = invoice.get("invoiceDate")
        if date_val:
            if isinstance(date_val, str):
                ws_invoices.cell(row=idx, column=9).value = date_val.split("T")[0]
            else:
                ws_invoices.cell(row=idx, column=9).value = date_val

        ws_invoices.cell(row=idx, column=10).value = invoice.get("category", "")

    # Add totals row
    total_row = len(invoices) + 2
    ws_invoices.cell(row=total_row, column=1).value = "TOTAL"
    ws_invoices.cell(row=total_row, column=1).font = Font(bold=True)

    # Sum formulas
    base_col = get_column_letter(5)
    tax_col = get_column_letter(6)
    total_col = get_column_letter(7)

    sum_base = ws_invoices.cell(row=total_row, column=5)
    sum_base.value = f"=SUM({base_col}2:{base_col}{len(invoices) + 1})"
    sum_base.font = Font(bold=True)
    sum_base.number_format = "#,##0.00"

    sum_tax = ws_invoices.cell(row=total_row, column=6)
    sum_tax.value = f"=SUM({tax_col}2:{tax_col}{len(invoices) + 1})"
    sum_tax.font = Font(bold=True)
    sum_tax.number_format = "#,##0.00"

    sum_total = ws_invoices.cell(row=total_row, column=7)
    sum_total.value = f"=SUM({total_col}2:{total_col}{len(invoices) + 1})"
    sum_total.font = Font(bold=True)
    sum_total.number_format = "#,##0.00"

    # Formatting
    for col_num in range(1, len(headers) + 1):
        ws_invoices.column_dimensions[get_column_letter(col_num)].width = 18

    if invoices:
        set_alternating_rows(ws_invoices, 2, len(invoices) + 1)
        add_thin_border(ws_invoices, 1, total_row)

    # Sheet 2: Resumen (Summary by Category)
    ws_summary = wb.create_sheet("Resumen", 1)
    summary_headers = ["Categoría", "Cantidad", "Base Imponible", "IVA", "Total"]
    add_header_row(ws_summary, summary_headers)

    summary_data = data.get("summary", [])
    for idx, item in enumerate(summary_data, start=2):
        ws_summary.cell(row=idx, column=1).value = item.get("category", "Sin categoría")
        ws_summary.cell(row=idx, column=2).value = item.get("count", 0)

        base_cell = ws_summary.cell(row=idx, column=3)
        base_cell.value = format_currency(item.get("sumBase"))
        base_cell.number_format = "#,##0.00"

        tax_cell = ws_summary.cell(row=idx, column=4)
        tax_cell.value = format_currency(item.get("sumTax"))
        tax_cell.number_format = "#,##0.00"

        total_cell = ws_summary.cell(row=idx, column=5)
        total_cell.value = format_currency(item.get("sumTotal"))
        total_cell.number_format = "#,##0.00"

    for col_num in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col_num)].width = 18

    if summary_data:
        set_alternating_rows(ws_summary, 2, len(summary_data) + 1)
        add_thin_border(ws_summary, 1, len(summary_data) + 1)

    return wb


def generate_emails_report(data):
    """Generate emails report with Emails and Estadísticas sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Emails
    ws_emails = wb.create_sheet("Emails", 0)
    headers = ["Fecha", "De", "Asunto", "Categoría", "Prioridad", "Leído"]
    add_header_row(ws_emails, headers)

    emails = data.get("emails", [])
    for idx, email in enumerate(emails, start=2):
        # Date
        date_val = email.get("date")
        if date_val:
            if isinstance(date_val, str):
                ws_emails.cell(row=idx, column=1).value = date_val.split("T")[0]
            else:
                ws_emails.cell(row=idx, column=1).value = date_val

        ws_emails.cell(row=idx, column=2).value = email.get("fromEmail", "")
        ws_emails.cell(row=idx, column=3).value = email.get("subject", "")
        ws_emails.cell(row=idx, column=4).value = email.get("category", "")
        ws_emails.cell(row=idx, column=5).value = email.get("priority", "")
        ws_emails.cell(row=idx, column=6).value = "Sí" if email.get("isRead") else "No"

    for col_num in range(1, len(headers) + 1):
        ws_emails.column_dimensions[get_column_letter(col_num)].width = 18

    if emails:
        set_alternating_rows(ws_emails, 2, len(emails) + 1)
        add_thin_border(ws_emails, 1, len(emails) + 1)

    # Sheet 2: Estadísticas
    ws_stats = wb.create_sheet("Estadísticas", 1)

    # Category stats
    ws_stats.cell(row=1, column=1).value = "Categoría"
    ws_stats.cell(row=1, column=2).value = "Cantidad"
    ws_stats.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    ws_stats.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")
    ws_stats.cell(row=1, column=1).fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
    ws_stats.cell(row=1, column=2).fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")

    category_stats = data.get("categoryStats", [])
    for idx, stat in enumerate(category_stats, start=2):
        ws_stats.cell(row=idx, column=1).value = stat.get("category", "Sin categoría")
        ws_stats.cell(row=idx, column=2).value = stat.get("count", 0)

    # Priority stats (right side)
    ws_stats.cell(row=1, column=4).value = "Prioridad"
    ws_stats.cell(row=1, column=5).value = "Cantidad"
    ws_stats.cell(row=1, column=4).font = Font(bold=True, color="FFFFFF")
    ws_stats.cell(row=1, column=5).font = Font(bold=True, color="FFFFFF")
    ws_stats.cell(row=1, column=4).fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
    ws_stats.cell(row=1, column=5).fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")

    priority_stats = data.get("priorityStats", [])
    for idx, stat in enumerate(priority_stats, start=2):
        ws_stats.cell(row=idx, column=4).value = stat.get("priority", "")
        ws_stats.cell(row=idx, column=5).value = stat.get("count", 0)

    ws_stats.column_dimensions["A"].width = 18
    ws_stats.column_dimensions["B"].width = 12
    ws_stats.column_dimensions["D"].width = 18
    ws_stats.column_dimensions["E"].width = 12

    return wb


def generate_executive_report(data):
    """Generate executive summary report."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Resumen Ejecutivo
    ws_exec = wb.create_sheet("Resumen Ejecutivo", 0)

    row = 1

    # Email Stats
    ws_exec.cell(row=row, column=1).value = "ESTADÍSTICAS DE EMAIL"
    ws_exec.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    ws_exec.cell(row=row, column=1).value = "Total de Emails"
    ws_exec.cell(row=row, column=2).value = data.get("totalEmails", 0)
    row += 1

    email_by_category = data.get("emailsByCategory", [])
    for stat in email_by_category:
        ws_exec.cell(row=row, column=1).value = f"  - {stat.get('category', 'Sin categoría')}"
        ws_exec.cell(row=row, column=2).value = stat.get("count", 0)
        row += 1

    row += 1
    ws_exec.cell(row=row, column=1).value = "ESTADÍSTICAS DE FACTURAS"
    ws_exec.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    ws_exec.cell(row=row, column=1).value = "Total Facturado"
    total_invoiced = ws_exec.cell(row=row, column=2)
    total_invoiced.value = format_currency(data.get("totalInvoiced"))
    total_invoiced.number_format = "#,##0.00"
    row += 1

    invoices_by_category = data.get("invoicesByCategory", [])
    for stat in invoices_by_category:
        ws_exec.cell(row=row, column=1).value = f"  - {stat.get('category', 'Sin categoría')}"
        total_cell = ws_exec.cell(row=row, column=2)
        total_cell.value = format_currency(stat.get("total"))
        total_cell.number_format = "#,##0.00"
        row += 1

    row += 1
    ws_exec.cell(row=row, column=1).value = "TOP 5 PROVEEDORES"
    ws_exec.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    ws_exec.cell(row=row, column=1).value = "Proveedor"
    ws_exec.cell(row=row, column=2).value = "Monto"
    ws_exec.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    ws_exec.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
    ws_exec.cell(row=row, column=1).fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
    ws_exec.cell(row=row, column=2).fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
    row += 1

    top_providers = data.get("topProviders", [])
    for provider in top_providers:
        ws_exec.cell(row=row, column=1).value = provider.get("issuerName", "")
        total_cell = ws_exec.cell(row=row, column=2)
        total_cell.value = format_currency(provider.get("total"))
        total_cell.number_format = "#,##0.00"
        row += 1

    ws_exec.column_dimensions["A"].width = 30
    ws_exec.column_dimensions["B"].width = 15

    # Sheet 2: Facturas (full invoice list)
    ws_invoices = wb.create_sheet("Facturas", 1)
    headers = ["Nº Factura", "Emisor", "NIF", "Concepto", "Base Imponible", "IVA", "Total", "Moneda", "Fecha", "Categoría"]
    add_header_row(ws_invoices, headers)

    invoices = data.get("invoices", [])
    for idx, invoice in enumerate(invoices, start=2):
        ws_invoices.cell(row=idx, column=1).value = invoice.get("invoiceNumber", "")
        ws_invoices.cell(row=idx, column=2).value = invoice.get("issuerName", "")
        ws_invoices.cell(row=idx, column=3).value = invoice.get("issuerNif", "")
        ws_invoices.cell(row=idx, column=4).value = invoice.get("concept", "")

        amount_cell = ws_invoices.cell(row=idx, column=5)
        amount_cell.value = format_currency(invoice.get("amount"))
        amount_cell.number_format = "#,##0.00"

        tax_cell = ws_invoices.cell(row=idx, column=6)
        tax_cell.value = format_currency(invoice.get("tax"))
        tax_cell.number_format = "#,##0.00"

        total_cell = ws_invoices.cell(row=idx, column=7)
        total_cell.value = format_currency(invoice.get("totalAmount"))
        total_cell.number_format = "#,##0.00"

        ws_invoices.cell(row=idx, column=8).value = invoice.get("currency", "EUR")

        date_val = invoice.get("invoiceDate")
        if date_val:
            if isinstance(date_val, str):
                ws_invoices.cell(row=idx, column=9).value = date_val.split("T")[0]
            else:
                ws_invoices.cell(row=idx, column=9).value = date_val

        ws_invoices.cell(row=idx, column=10).value = invoice.get("category", "")

    for col_num in range(1, len(headers) + 1):
        ws_invoices.column_dimensions[get_column_letter(col_num)].width = 18

    if invoices:
        set_alternating_rows(ws_invoices, 2, len(invoices) + 1)
        add_thin_border(ws_invoices, 1, len(invoices) + 1)

    # Sheet 3: Actividad Email
    ws_activity = wb.create_sheet("Actividad Email", 2)
    activity_headers = ["Fecha", "De", "Asunto", "Categoría", "Prioridad", "Leído"]
    add_header_row(ws_activity, activity_headers)

    emails = data.get("emails", [])
    for idx, email in enumerate(emails, start=2):
        date_val = email.get("date")
        if date_val:
            if isinstance(date_val, str):
                ws_activity.cell(row=idx, column=1).value = date_val.split("T")[0]
            else:
                ws_activity.cell(row=idx, column=1).value = date_val

        ws_activity.cell(row=idx, column=2).value = email.get("fromEmail", "")
        ws_activity.cell(row=idx, column=3).value = email.get("subject", "")
        ws_activity.cell(row=idx, column=4).value = email.get("category", "")
        ws_activity.cell(row=idx, column=5).value = email.get("priority", "")
        ws_activity.cell(row=idx, column=6).value = "Sí" if email.get("isRead") else "No"

    for col_num in range(1, len(activity_headers) + 1):
        ws_activity.column_dimensions[get_column_letter(col_num)].width = 18

    if emails:
        set_alternating_rows(ws_activity, 2, len(emails) + 1)
        add_thin_border(ws_activity, 1, len(emails) + 1)

    return wb


def generate_expenses_report(data):
    """Generate expenses report grouped by issuer and category."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Gastos Recurrentes
    ws_recurring = wb.create_sheet("Gastos Recurrentes", 0)
    headers = ["Proveedor", "Última Factura", "Frecuencia", "Monto Medio", "Total Anual (Est.)"]
    add_header_row(ws_recurring, headers)

    recurring = data.get("recurringExpenses", [])
    for idx, item in enumerate(recurring, start=2):
        ws_recurring.cell(row=idx, column=1).value = item.get("issuer", "")

        date_val = item.get("lastInvoiceDate")
        if date_val:
            if isinstance(date_val, str):
                ws_recurring.cell(row=idx, column=2).value = date_val.split("T")[0]
            else:
                ws_recurring.cell(row=idx, column=2).value = date_val

        ws_recurring.cell(row=idx, column=3).value = item.get("frequency", "")

        avg_cell = ws_recurring.cell(row=idx, column=4)
        avg_cell.value = format_currency(item.get("avgAmount"))
        avg_cell.number_format = "#,##0.00"

        annual_cell = ws_recurring.cell(row=idx, column=5)
        annual_cell.value = format_currency(item.get("estimatedAnnual"))
        annual_cell.number_format = "#,##0.00"

    for col_num in range(1, len(headers) + 1):
        ws_recurring.column_dimensions[get_column_letter(col_num)].width = 20

    if recurring:
        set_alternating_rows(ws_recurring, 2, len(recurring) + 1)
        add_thin_border(ws_recurring, 1, len(recurring) + 1)

    # Sheet 2: Por Categoría
    ws_by_category = wb.create_sheet("Por Categoría", 1)
    category_headers = ["Categoría", "Cantidad", "Monto Total", "Monto Medio"]
    add_header_row(ws_by_category, category_headers)

    by_category = data.get("expensesByCategory", [])
    for idx, item in enumerate(by_category, start=2):
        ws_by_category.cell(row=idx, column=1).value = item.get("category", "")
        ws_by_category.cell(row=idx, column=2).value = item.get("count", 0)

        total_cell = ws_by_category.cell(row=idx, column=3)
        total_cell.value = format_currency(item.get("total"))
        total_cell.number_format = "#,##0.00"

        avg_cell = ws_by_category.cell(row=idx, column=4)
        avg_cell.value = format_currency(item.get("average"))
        avg_cell.number_format = "#,##0.00"

    for col_num in range(1, len(category_headers) + 1):
        ws_by_category.column_dimensions[get_column_letter(col_num)].width = 20

    if by_category:
        set_alternating_rows(ws_by_category, 2, len(by_category) + 1)
        add_thin_border(ws_by_category, 1, len(by_category) + 1)

    # Sheet 3: Mensual
    ws_monthly = wb.create_sheet("Mensual", 2)
    monthly_headers = ["Mes", "Cantidad", "Monto Total"]
    add_header_row(ws_monthly, monthly_headers)

    monthly = data.get("monthlyExpenses", [])
    for idx, item in enumerate(monthly, start=2):
        ws_monthly.cell(row=idx, column=1).value = item.get("month", "")
        ws_monthly.cell(row=idx, column=2).value = item.get("count", 0)

        total_cell = ws_monthly.cell(row=idx, column=3)
        total_cell.value = format_currency(item.get("total"))
        total_cell.number_format = "#,##0.00"

    for col_num in range(1, len(monthly_headers) + 1):
        ws_monthly.column_dimensions[get_column_letter(col_num)].width = 20

    if monthly:
        set_alternating_rows(ws_monthly, 2, len(monthly) + 1)
        add_thin_border(ws_monthly, 1, len(monthly) + 1)

    return wb


def main():
    """Main entry point."""
    try:
        # Read JSON from stdin
        json_input = sys.stdin.read()
        data = json.loads(json_input)

        report_type = data.get("type", "invoices")

        # Generate appropriate workbook
        if report_type == "invoices":
            wb = generate_invoices_report(data)
        elif report_type == "emails":
            wb = generate_emails_report(data)
        elif report_type == "executive":
            wb = generate_executive_report(data)
        elif report_type == "expenses":
            wb = generate_expenses_report(data)
        else:
            raise ValueError(f"Unknown report type: {report_type}")

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)

        # Output the file path to stdout
        print(tmp_path)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
